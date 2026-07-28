# -*- coding: utf-8 -*-
"""从《单身宿舍房间安排明细及排查表2026（新）.xls》导入宿舍床位数据到 dorm 表，
并生成一份干净的名册 xlsx（按地区/房号/床位排序 + 空床位汇总）。

用法：  python tools/seed_dorm.py
依赖：  xlrd（读旧版 .xls）、openpyxl（写 .xlsx）
"""
import os
import sys
from datetime import datetime, timedelta

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE)
import db  # noqa: E402

# 源表（相对本仓库定位到同级的 qi-bangong）
SRC = os.path.join(
    BASE, "..", "qi-bangong", "七星", "职工宿舍&平房人员",
    "单身宿舍房间安排明细及排查表2026（新）.xls",
)
OUT_XLSX = os.path.join(BASE, "单身宿舍名册（整理版）.xlsx")

# 各点位床位容量（依据源表首行备注登记的床位数）
REGION_CAP = [
    ("望京经干院", 25),
    ("西站中雅大厦", 12),
    ("望京南湖中园", 5),
    ("芳群园三区15号楼", 3),
    ("芳古园一区14号楼", 3),
    ("芳群园四区1号楼", 3),
    ("定安东里6号楼", 3),
]
REGION_ORDER = {name: i for i, (name, _) in enumerate(REGION_CAP)}


def _xldate(v):
    try:
        v = float(v)
        if v > 0:
            return (datetime(1899, 12, 30) + timedelta(days=v)).strftime("%Y-%m-%d")
    except (TypeError, ValueError):
        pass
    return ""


def parse():
    import xlrd
    wb = xlrd.open_workbook(SRC)
    sh = wb.sheet_by_index(0)

    def cv(r, c):
        v = sh.cell_value(r, c)
        if isinstance(v, float) and v == int(v):
            v = int(v)
        return str(v).strip()

    def maybe_date(r, c):
        """数值当作 Excel 日期序列转换，否则原样返回文本。"""
        v = sh.cell_value(r, c)
        if isinstance(v, (int, float)) and v > 20000:  # 序列日期
            return _xldate(v)
        return cv(r, c)

    recs = []
    for r in range(2, sh.nrows):
        if not (cv(r, 1).isdigit() and cv(r, 2)):
            continue  # 只取有序号+姓名的主记录行
        note_raw = cv(r, 17)
        sign = cv(r, 18)
        code = next((t for t in note_raw.split() if t.startswith("CESI-DSSS")), "")
        note_txt = " ".join(t for t in note_raw.split() if not t.startswith("CESI-DSSS")).strip()
        note_full = " ".join(x for x in (note_txt, sign) if x).strip()

        if "搬出" in note_full or "搬走" in note_full:
            status = "已搬出"
        elif "人才宿舍" in note_full or "人才公寓" in note_full:
            status = "人才公寓"
        else:
            status = "在住"

        recs.append({
            "region": cv(r, 4), "room_no": cv(r, 5), "bed_no": cv(r, 6),
            "gender": cv(r, 7), "name": cv(r, 2), "dept": cv(r, 3),
            "phone": cv(r, 8), "move_in": _xldate(sh.cell_value(r, 9)),
            "adjust_date": maybe_date(r, 10), "fee_tier": "",  # 源表勾选项不可机读，留空待人工确认
            "status": status, "code": code, "notes": note_full,
        })
    return recs


def _sort_key(x):
    return (REGION_ORDER.get(x["region"], 99), x["room_no"], x["bed_no"])


def import_db(recs):
    db.init_db()
    db.run("DELETE FROM dorm")
    cols = ["region", "room_no", "bed_no", "gender", "name", "dept", "phone",
            "move_in", "adjust_date", "fee_tier", "status", "code", "notes"]
    for x in sorted(recs, key=_sort_key):
        ph = ",".join("?" * len(cols))
        db.run(f"INSERT INTO dorm({','.join(cols)}) VALUES({ph})", [x[c] for c in cols])
    print(f"已写入 dorm 表：{len(recs)} 条")


def make_xlsx(recs):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    hdr_fill = PatternFill("solid", fgColor="1F3B57")
    hdr_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "宿舍名册"
    headers = ["地区", "房号", "床位", "男女", "姓名", "部门", "联系电话",
               "入住时间", "状态", "备案编号", "备注"]
    ws.append(headers)
    for r in sorted(recs, key=_sort_key):
        ws.append([r["region"], r["room_no"], r["bed_no"], r["gender"], r["name"],
                   r["dept"], r["phone"], r["move_in"], r["status"], r["code"], r["notes"]])
    widths = [15, 10, 6, 8, 9, 22, 14, 12, 10, 20, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill, cell.font, cell.alignment, cell.border = hdr_fill, hdr_font, center, border
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).border = border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # 空床位汇总
    ws2 = wb.create_sheet("空床位汇总")
    ws2.append(["宿舍地区", "床位容量", "占用", "空床位"])
    # 占用含在住与人才公寓，仅排除已搬出
    active = [x for x in recs if x["status"] != "已搬出"]
    tot_cap = tot_occ = 0
    for name, cap in REGION_CAP:
        occ = sum(1 for x in active if x["region"] == name)
        tot_cap += cap
        tot_occ += occ
        ws2.append([name, cap, occ, cap - occ])
    ws2.append(["合计", tot_cap, tot_occ, tot_cap - tot_occ])
    for i, w in enumerate([18, 10, 8, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for c in range(1, 5):
        cell = ws2.cell(1, c)
        cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, center
    for c in range(1, 5):
        f = ws2.cell(ws2.max_row, c)
        f.font = Font(bold=True)
    wb.save(OUT_XLSX)
    print(f"已生成名册：{OUT_XLSX}")


if __name__ == "__main__":
    data = parse()
    import_db(data)
    make_xlsx(data)
    # 概览
    from collections import Counter
    print("状态分布：", dict(Counter(x["status"] for x in data)))
